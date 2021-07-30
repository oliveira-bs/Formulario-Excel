import codecs
from openpyxl import load_workbook
import os
import chardet
from tkinter import *
import tkinter as tk 
from tkinter import messagebox
from tkinter import filedialog
from tkinter import ttk
from tkinter.filedialog import askopenfilename


def check_folder_ppf(): #Checa se existe o diretorio 'PPF'
    diretorio=str('PPF')
    c_f_ppf= os.path.exists('./{}'.format(diretorio))
    
    if(c_f_ppf == True):
        return True
    else:
        return False
    
def make_dir_ppf(): #Cria um diretorio chamado 'PPF'
    diretorio=str('PPF')
    c_f_ppf= os.path.exists('./{}'.format(diretorio))
    
    if(c_f_ppf == True):
        pass
    else:
        os.makedirs("./{}".format(diretorio))
        
def make_dir_ficha(): #Cria um diretorio chamado 'Ficha_Tratamento'
    diretorio=str('Ficha_Tratamento')
    c_f_ppf= os.path.exists('./{}'.format(diretorio))
    
    if(c_f_ppf == True):
        pass
    else:
        os.makedirs("./{}".format(diretorio))

def nome_arq(arq1):
    
    caminho1 = "/" in arq1
    caminho2 = "\\" in arq1
    if(caminho1) == True:
        arq_caminho = arq1.split("/")
        arq_caminho = arq_caminho[-1]
        arq_caminho = arq_caminho.split("'")
        arq_caminho = arq_caminho[0]        
        arq1 = arq_caminho.strip()

    elif(caminho2) == True:
        arq_caminho = arq1.split("\\")
        arq_caminho = arq_caminho[-1]
        arq_caminho = arq_caminho.split("'")
        arq_caminho = arq_caminho[0]          
        arq1 = arq_caminho.strip()

    else:
        arq1 = arq1
        
    extension = '.PPF' in arq1
    
    if(extension) == True:
        name = arq1
    
    else:
        name = "{}.PPF".format(arq1)

    ppf= os.path.exists('./PPF/{}'.format(name))
 
    if(ppf == True):
        name = name
        
    else:
        name = "Erro"

    return name

boost = 0
lista_arquivo = []
def window_main(lista_arquivo):
    window = Tk()
    window.title('Automatizador de ficha Plan_Onco')
    window.geometry('380x200')

    lmain = Label(window, text= """Escolha o padrão da
ficha de tratamento?""", font= ('',8))
    lmain.grid(row= 0, pady= 10, sticky= EW)

    def close_window():
        window.destroy()

    def OpenFile():
        global lista_arquivo
        name = askopenfilename(initialdir="./PPF/",
                            filetypes =(("PPF Files", ".ppf .PPF"),("All Files","*.*")),
                            title = "Choose a file.", multiple= True)
        try:
            for arq in name:
                lista_arquivo.append(arq)
        except:
            messagebox.showinfo("Erro", message= "Erro ao adicionar os arquivos")
        if len(lista_arquivo)>0:
            Label(window, text= """Planejamento e arquivos 
selecionados""", 
            font= ('',8)).grid(pady= 20,row= 4, column= 0) 
            Button(window, command= close_window, text= 'Avançar', width= 15, 
            font= ('',9)).grid(pady= 20,row= 4, column= 1) 

    def boost_concomitante(value):
        global boost
        boost = value

    xdist_button = 25
    ydist_button = 1
    width_button = 20   
    b1 = Button(window, command= lambda *args: [boost_concomitante(1), OpenFile()], 
                text= '6 CAMPOS', width= width_button, font= ('',8))
    b1.grid(pady= ydist_button, padx= xdist_button, row= 1, column= 0 )

    b2 = Button(window, command= lambda *args: [boost_concomitante(3), OpenFile()], 
                text= '8 CAMPOS', width= width_button, font= ('',8))
    b2.grid(pady= ydist_button, padx= xdist_button, row= 2, column= 0)
    b3 = Button(window, command= lambda *args: [boost_concomitante(3), OpenFile()] , 
                text= 'MAMA+SUPRA+BC', width= width_button, font= ('',8))
    b3.grid(pady= ydist_button, padx= xdist_button, row= 3, column= 0)
    b4 = Button(window, command= lambda *args: [boost_concomitante(4), OpenFile()] , 
                text= 'MAMA(4)+BC', width= width_button, font= ('',8))
    b4.grid(pady= ydist_button, padx= xdist_button, row= 1, column= 1)
    b5 = Button(window, command= lambda *args: [boost_concomitante(5), OpenFile()] , 
                text= 'MAMA(2)+BC', width= width_button, font= ('',8))
    b5.grid(pady= ydist_button, padx= xdist_button, row= 2, column= 1)


    window.mainloop()    
    return (boost, lista_arquivo)

def window_boost():
    window2 = Tk()
    window2.title('Dose Boost')
    window2.geometry('350x100')
    l3 = Label(window2, text="""Digite a dose inicial 
    para o boost concomitante""", font= ('',8))
    l3.grid(row= 0, column= 0, padx= 10, pady= 12)
    boost_dose = IntVar()
    t3 = Entry(window2, textvariable= boost_dose, font= ('',10))
    t3.grid(row= 0, column= 1, padx= 0, pady= 10)
    
    def boost():
        x = boost_dose.get()
        return x
    
    def destroy():
        window2.destroy()

    b3_ok = Button(window2, command= lambda: [boost(), destroy()], text= 'Aplicar', font= ('',8))
    b3_ok.grid(row= 1, column= 1, ipadx= 20, padx= 0, pady= 0)
    
    window2.mainloop()
    x= boost()
    return (x)

sort_path = []
def sort_ppf(lista_arquivo):
    window3 = Tk()
    window3.title('Teste Main PPF')
    window3.geometry('450x150')

    Label(window3, text= """Escolha o arquivo PPF padrão""", font= ('',8)).pack(padx=5, pady= 5)

    def close_window3():
        window3.destroy()

    def boost_concomitante(value, lista_arquivo):
        global sort_path
        choice = value
        arquivo = lista_arquivo[choice]
        lista_arquivo.pop(choice)
        lista_arquivo.append(arquivo)
        sort_path = lista_arquivo
         
    width_button = 40

    for i in lista_arquivo:
        filename = nome_arq(i)  
        Button(window3, command= lambda i=i: [boost_concomitante(lista_arquivo.index(i), lista_arquivo), 
        close_window3()], text= filename, width= width_button, 
        font= ('',8)).pack(pady= 1)
    window3.mainloop()    
    return sort_path

class App():  
    def __init__(self):
        self.window3 = Tk()
        self.window3.geometry('250x200')
        self.window3.title('As hswjsdhj')
        Label(self.window3, text ='Botão fechar', font= ('',8)).pack(padx=5, pady= 5)

        for i in range(0,5):
            Button(self.window3, text = 'root quit', command= self.quit, width= 30, font= ('',8)).pack(pady= 1)

        self.window3.mainloop()
        
    def quit(self):
        self.window3.destroy()

def dose(arq1):
    mestre="[Header]"
    parametro="TDOSE"
    with open(arq1, 'rb') as rawdata:
        result = chardet.detect(rawdata.read())

    with codecs.open(arq1,'r',encoding=result['encoding'], errors='ignore') as arq:
        for linha in arq:
            linha=linha.strip()     #limpa os espaços externos da linha percorrida para a busca do termo mestre
          
            lista=linha.split("=")  #separa os elementos da linha percorrida em lista durante a busca do termo mestre
            if(lista[0]==mestre):

                for linha in arq:
                    linha=linha.strip()     #limpa os espaços externos da linha percorrida para a busca do termo parametro
                    lista=linha.split("=")  #separa os elementos da linha percorrida em lista durante a busca do termo parametro
                    listaf=[]

                    for l in lista:
                        clean=l.strip()     #limpa os espaços externos de cada elemento da lista durante a busca do termo parametro
                        listaf.append(clean)
                         
                    if(listaf[0]==parametro):           
                        dose=float(listaf[1])   #transforma o elemento string da listaf em numero real(float)
                        return dose
                
def aplicacoes(arq1):
    mestre="[Header]"
    parametro="NUMFRACTIONS"
    with open(arq1, 'rb') as rawdata:
        result = chardet.detect(rawdata.read())
    with codecs.open(arq1,'r',encoding=result['encoding'], errors='ignore') as arq:
        for linha in arq:
            linha=linha.strip()     #limpa os espaços externos da linha percorrida para a busca do termo mestre
            lista=linha.split("=")  #separa os elementos da linha percorrida em lista durante a busca do termo mestre
            
            if(lista[0]==mestre):
                for linha in arq:
                    linha=linha.strip()     #limpa os espaços externos da linha percorrida para a busca do termo parametro
                    lista=linha.split("=")  #separa os elementos da linha percorrida em lista durante a busca do termo parametro
                    listaf=[]
                    
                    for l in lista:
                        clean=l.strip()     #limpa os espaços externos de cada elemento da lista durante a busca do termo parametro
                        listaf.append(clean)
                    if(listaf[0]==parametro):           
                        aplicacoes=int(listaf[1])   #transforma o elemento string da listaf em numero real(float)
                        return aplicacoes
            
def campo_inicial(arq1):
    while True:
        mestre="[FIELD_"
        with open(arq1, 'rb') as rawdata:
            result = chardet.detect(rawdata.read())

        with codecs.open(arq1,'r',encoding=result['encoding'], errors='ignore') as arq:
            for linha in arq:
                linha=linha.strip()     #limpa os espaços externos da linha percorrida para a busca do termo mestre

                if(mestre in linha):
                    campo_1 = arq.readline()
                    limpeza = campo_1.strip()
                    primeiro_campo = int(limpeza)                    
                    break
        return primeiro_campo

def quant_campos(arq1):
    mestre="[Header]"
    parametro="FIELDS"
    with open(arq1, 'rb') as rawdata:
        result = chardet.detect(rawdata.read())
    
    with codecs.open(arq1,'r',encoding=result['encoding'], errors='ignore') as arq:    
        for linha in arq:
            linha=linha.strip()     #limpa os espaços externos da linha percorrida para a busca do termo mestre          
            lista=linha.split("=")  #separa os elementos da linha percorrida em lista durante a busca do termo mestre

            if(lista[0]==mestre):
                for linha in arq:
                    linha=linha.strip()     #limpa os espaços externos da linha percorrida para a busca do termo parametro
                    lista=linha.split("=")  #separa os elementos da linha percorrida em lista durante a busca do termo parametro
                    listaf=[]
                    for l in lista:
                        clean=l.strip()     #limpa os espaços externos de cada elemento da lista durante a busca do termo parametro
                        listaf.append(clean)

                    if(listaf[0]==parametro):           
                        total_campos = int(listaf[1])#transforma o elemento string da listaf em numero real(float)
                        return total_campos

def list_campos(arq1):
    mestre="[FIELD_"
    fields = list()
    with open(arq1, 'rb') as rawdata:
        result = chardet.detect(rawdata.read())
    
    with codecs.open(arq1,'r',encoding=result['encoding'], errors='ignore') as arq:
        for linha in arq:        
            linha=linha.strip()     #limpa os espaços externos da linha percorrida para a busca do termo mestre
   
            if(mestre in linha):
                for i in range(0,1):#Lê as linhas até chegar no valor para o tamanho de campo Y
                    field =arq.readline()
                    field = int(field)
                    fields.append(field)
        return (fields)

def conf_campos(arq1, campo):
    mestre="[FIELD_{}]".format(campo)

    with open(arq1, 'rb') as rawdata:
        result = chardet.detect(rawdata.read())
    
    with codecs.open(arq1,'r',encoding=result['encoding'], errors='ignore') as arq:
        for linha in arq:
            linha=linha.strip()     #limpa os espaços externos da linha percorrida para a busca do termo mestre
            lista=linha.split("=")  #separa os elementos da linha percorrida em lista durante a busca do termo mestre

            if(lista[0]==mestre):
                for i in range(0,3):    #Lê as linhas até chegar no valor para o tamanho de campo Y
                    Y=arq.readline()
                Y= float(Y)
                Y= Y/10
                Y= round(Y,1)
                X=arq.readline()        #Lê as linhas até chegar no valor para o tamanho de campo X
                X= float(X)
                X= X/10
                X= round(X,1)

                for l in range(0,2):    #Lê as linhas até chegar no valor Angulo do Gantry
                    ang_gantry=arq.readline()
                ang_gantry=float(ang_gantry)
                ang_gantry=int(ang_gantry)
                
                for l in range(0,2):    #Lê as linhas até chegar no valor Angulo do Colimador
                    col_ang=arq.readline()
                col_ang=float(col_ang)
                col_ang=int(col_ang)
                
                for l in range(0,2):    #Lê as linhas até chegar na strinf Wedge Filt
                    fwedge=(arq.readline()).strip()
                pwedge=(arq.readline()).strip()     #Lê as linhas até chegar na strinf Wedge Pos

                
                for l in range(0,3):    #Lê as linhas até chegar no valor para o tamanho de campo Y2
                    Y2=arq.readline()
                Y2=float(Y2)
                Y1=arq.readline()       #Lê as linhas até chegar no valor para o tamanho de campo Y1
                Y1=float(Y1)
                X1=arq.readline()       #Lê as linhas até chegar no valor para o tamanho de campo X1
                X1=float(X1)
                X2=arq.readline()       #Lê as linhas até chegar no valor para o tamanho de campo X2
                X2=float(X2)
                SSD=arq.readline()      #Lê as linhas até chegar no valor para o tamanho de campo SSD
                SSD=float(SSD)
                SSD=(SSD/10)
                
                for l in range(0,3):#Lê as linhas até chegar no valor da Unidade Minitora
                    mon_units=arq.readline()
                mon_units=float(mon_units)
                  
    return [Y, X, ang_gantry, col_ang, fwedge, pwedge, Y2, Y1, X1, X2, SSD, mon_units]
                
def setup1(arq1, primeiro_campo, total_campos, planilha):
    pc= planilha.cell
    merge = planilha.merge_cells
    pc(1, 4, value=dose(arq1))
    pc(1, 7, value=aplicacoes(arq1))
    campos = list_campos(arq1)
    for l in campos:
        dados_campo = conf_campos(arq1,l)
        dc = dados_campo
        
        if(l<3):
            pc(3,(8+(l-1)*2)).value=dc[-2]  #Add Distancia fonte pele
            pc(9,(8+(l-1)*2)).value=dc[2]   #Add Angulo do Gantry
            pc(10,(8+(l-1)*2)).value=dc[3]  #Add Angulo do Colimador
            if(dc[4]!="NONE"):
                pc(11,(8+(l-1)*2)).value="{} {}".format(dc[4], dc[5])   #Add Tipo e Posição do Filtro Fisico
            else:
                pc(11,(8+(l-1)*2),value="")     #Add Tipo e Posição do Filtro Fisico
            pc(13,(8+(l-1)*2)).value=dc[-1]     #Add Unidade Monitora
        else:
            pc(3,(9+(l-1)*2)).value=dc[-2]      #Add Distancia fonte pele
            pc(9,(9+(l-1)*2)).value=dc[2]       #Add Angulo do Gantry
            pc(10,(9+(l-1)*2)).value=dc[3]      #Add Angulo do Colimador
            if(dc[4]!="NONE"):
                pc(11,(9+(l-1)*2)).value="{} {}".format(dc[4], dc[5])   #Add Tipo e Posição do Filtro Fisico
            else:
                pc(11,(9+(l-1)*2),value="")     #Add Tipo e Posição do Filtro Fisico
            pc(13,(9+(l-1)*2)).value=dc[-1]     #Add Unidade Monitora
        
        #Condiçoes e Coordenadas para y, X, Y2, Y1, X1, X2
                        
        if((dc[6] or dc[7])!=0 and (dc[8] or dc[9])!=0):
            if(l<3):
                if(dc[-5] == dc[-6]):
                    merge(start_row = 5, start_column = 9+(l-1)*2, end_row = 6, end_column = 9+(l-1)*2)
                    cell = pc(5,(9+(l-1)*2))
                    cell.value=(dc[-5]/10)+(dc[-6]/10)  #Y1+Y2
                else:
                    pc(5,(9+(l-1)*2)).value=dc[-5]/10   #Y1
                    pc(6,(9+(l-1)*2)).value=dc[-6]/10   #Y2
                
                if(dc[-3] == dc[-4]):
                    merge(start_row = 7, start_column = 9+(l-1)*2, end_row = 8, end_column = 9+(l-1)*2)
                    cell = pc(7,(9+(l-1)*2))
                    cell.value=(dc[-3]/10)+(dc[-4]/10)  #X1+X2
                else:
                    pc(7,(9+(l-1)*2)).value=dc[-4]/10   #X1
                    pc(8,(9+(l-1)*2)).value=dc[-3]/10   #X2
            else:
                if(dc[-5] == dc[-6]):
                    merge(start_row = 5, start_column = 10+(l-1)*2, end_row = 6, end_column = 10+(l-1)*2)
                    cell = pc(5,(10+(l-1)*2))
                    cell.value=(dc[-5]/10)+(dc[-6]/10)  #Y1+Y2                
                else:
                    pc(5,(10+(l-1)*2)).value=dc[-5]/10  #Y1
                    pc(6,(10+(l-1)*2)).value=dc[-6]/10  #Y2
                    
                if(dc[-3] == dc[-4]):
                    merge(start_row = 7, start_column = 10+(l-1)*2, end_row = 8, end_column = 10+(l-1)*2)
                    cell = pc(7,(10+(l-1)*2))
                    cell.value=(dc[-3]/10)+(dc[-4]/10)  #X1+X2    
                else:
                    pc(7,(10+(l-1)*2)).value=dc[-4]/10  #X1
                    pc(8,(10+(l-1)*2)).value=dc[-3]/10  #X2
        elif((dc[6] and dc[7])==0 and (dc[8] or dc[9])!=0): #Y2 e  Y1 == 0
            Y2=dc[0]
            Y1=Y2
            if(l<3):   
                merge(start_row = 5, start_column = 9+(l-1)*2, end_row = 6, end_column = 9+(l-1)*2)
                cell = pc(5,(9+(l-1)*2))
                cell.value=Y2   #Y1=Y2=Y                        

                if(dc[-3] == dc[-4]):
                    merge(start_row = 7, start_column = 9+(l-1)*2, end_row = 8, end_column = 9+(l-1)*2)
                    cell = pc(7,(9+(l-1)*2))
                    cell.value=(dc[-3]/10)+(dc[-4]/10)  #X1+X2
                else:
                    pc(7,(9+(l-1)*2)).value=dc[-4]/10   #X1
                    pc(8,(9+(l-1)*2)).value=dc[-3]/10   #X2

            else:
                merge(start_row = 5, start_column = 10+(l-1)*2, end_row = 6, end_column = 10+(l-1)*2)
                cell = pc(5,(10+(l-1)*2))
                cell.value=Y2                
                
                if(dc[-3] == dc[-4]):
                    merge(start_row = 7, start_column = 10+(l-1)*2, end_row = 8, end_column = 10+(l-1)*2)
                    cell = pc(7,(10+(l-1)*2))
                    cell.value=(dc[-3]/10)+(dc[-4]/10)  #X1+X2

                else:
                    pc(7,(10+(l-1)*2)).value=dc[-4]/10  #X1
                    pc(8,(10+(l-1)*2)).value=dc[-3]/10  #X2
        
        elif((dc[8] and dc[9])==0 and (dc[6] or dc[7])!=0): #X2 e X1 == 0
            X2=dc[1]
            X1=X2
            if(l<3):
                merge(start_row = 7, start_column = 9+(l-1)*2, end_row = 8, end_column = 9+(l-1)*2)
                cell = pc(7,(9+(l-1)*2))
                cell.value=X2   #X1=X2 merge 

                if(dc[-5] == dc[-6]):
                    merge(start_row = 5, start_column = 9+(l-1)*2, end_row = 6, end_column = 9+(l-1)*2)
                    cell = pc(5,(9+(l-1)*2))
                    cell.value=2*(dc[-5]/10)    #Y1=Y2 merge
                else:
                    pc(5,(9+(l-1)*2)).value=dc[-5]/10   #Y1
                    pc(6,(9+(l-1)*2)).value=dc[-6]/10   #Y2                

            else:
                merge(start_row = 7, start_column = 10+(l-1)*2, end_row = 8, end_column = 10+(l-1)*2)
                cell = pc(7,(10+(l-1)*2))
                cell.value=X2   #X1=X2 merge 

                if(dc[-5] == dc[-6]):
                    merge(start_row = 5, start_column = 9+(l-1)*2, end_row = 6, end_column = 9+(l-1)*2)
                    cell = pc(5,(9+(l-1)*2))
                    cell.value=2*(dc[-5]/10)    #Y1=Y2 merge
                else:
                    pc(5,(10+(l-1)*2)).value=dc[-5]/10  #Y1
                    pc(6,(10+(l-1)*2)).value=dc[-6]/10  #Y2

        elif(dc[6]==0 and dc[7]==0 and dc[8]==0 and dc[9]==0):
            if(l<3):
                pc(4,(8+(l-1)*2)).value="{} x {}".format(dc[0],dc[1])   #Y x X
            else:
                pc(4,(9+(l-1)*2)).value="{} x {}".format(dc[0],dc[1])   #Y e X1
                
def setup2(arq1, dose_boost, primeiro_campo, total_campos, planilha):
    pc = planilha.cell
    merge = planilha.merge_cells
    pc(1, 4, value=dose(arq1))
    pc(1, 7, value=aplicacoes(arq1))
    pc(15,19).value = dose_boost    #Add dose incial do Boost Concomitante
    
    for l in range(primeiro_campo, primeiro_campo+total_campos):
        dados_campo = conf_campos(arq1,l)
        dc = dados_campo
        
        if(l<3):
            fc = 8      #fc: fator coluna = parametro inicial que indica a posição da coluna alvo
            pc(3,(fc+(l-1)*2)).value=dc[-2]     #Add Distancia fonte pele
            pc(9,(fc+(l-1)*2)).value=dc[2]      #Add Angulo do Gantry
            pc(10,(fc+(l-1)*2)).value=dc[3]     #Add Angulo do Colimador
            pc(13,(fc+(l-1)*2)).value=dc[-1]    #Add Unidade Monitora
            if(dc[4]!="NONE"):
                pc(11,(fc+(l-1)*2)).value="{} {}".format(dc[4], dc[5])  #Add Tipo e Posição do Filtro Fisico
            else:
                pc(11,(fc+(l-1)*2),value="")    #Add Tipo e Posição do Filtro Fisico
            
        elif(l<6):
            fc = 9      #fc: fator coluna = parametro inicial que indica a posição da coluna alvo
            pc(3,(fc+(l-1)*2)).value=dc[-2]     #Add Distancia fonte pele
            pc(9,(fc+(l-1)*2)).value=dc[2]      #Add Angulo do Gantry
            pc(10,(fc+(l-1)*2)).value=dc[3]     #Add Angulo do Colimador
            pc(13,(fc+(l-1)*2)).value=dc[-1]    #Add Unidade Monitora
            if(dc[4]!="NONE"):
                pc(11,(fc+(l-1)*2)).value="{} {}".format(dc[4], dc[5])  #Add Tipo e Posição do Filtro Fisico
            else:
                pc(11,(fc+(l-1)*2),value="")    #Add Tipo e Posição do Filtro Fisico
        
        else:
            fc = 11     #fc: fator coluna = parametro inicial que indica a posição da coluna alvo
            pc(3,(fc+(l-1)*2)).value=dc[-2]     #Add Distancia fonte pele
            pc(9,(fc+(l-1)*2)).value=dc[2]      #Add Angulo do Gantry
            pc(10,(fc+(l-1)*2)).value=dc[3]     #Add Angulo do Colimador
            pc(13,(fc+(l-1)*2)).value=dc[-1]    #Add Unidade Monitora
            if(dc[4]!="NONE"):
                pc(11,(fc+(l-1)*2)).value="{} {}".format(dc[4], dc[5])  #Add Tipo e Posição do Filtro Fisico
            else:
                pc(11,(fc+(l-1)*2),value="")    #Add Tipo e Posição do Filtro Fisico                
        
        #Condiçoes e Coordenadas para Y, X, Y2, Y1, X1, X2
                        
        if((dc[6] or dc[7])!=0 and (dc[8] or dc[9])!=0):
            if(l<3):
                fc = 9      #fc: fator coluna = parametro inicial que indica a posição da coluna alvo
                
                if(dc[-5] == dc[-6]):   #Y1 == Y2
                    merge(start_row = 5, start_column = fc+(l-1)*2, end_row = 6, end_column = fc+(l-1)*2)
                    cell = pc(5,(fc+(l-1)*2))
                    cell.value=2*(dc[-5]/10)    #Y1+Y2 
                else:
                    pc(5,(fc+(l-1)*2)).value=dc[-5]/10  #Y1
                    pc(6,(fc+(l-1)*2)).value=dc[-6]/10  #Y2
                
                if(dc[-4] == dc[-3]):   #X1 = X2
                    merge(start_row = 7, start_column = fc+(l-1)*2, end_row = 8, end_column = fc+(l-1)*2)
                    cell = pc(5,(fc+(l-1)*2))
                    cell.value=2*(dc[-4]/10)    #X1+X2 
                else:
                    pc(7,(fc+(l-1)*2)).value=dc[-4]/10  #X1
                    pc(8,(fc+(l-1)*2)).value=dc[-3]/10  #X2
                
            elif(l<6):
                fc = 10     #fc: fator coluna = parametro inicial que indica a posição da coluna alvo
                
                if(dc[-5] == dc[-6]):   #Y1 == Y2
                    merge(start_row = 5, start_column = fc+(l-1)*2, end_row = 6, end_column = fc+(l-1)*2)
                    cell = pc(5,(fc+(l-1)*2))
                    cell.value=2*(dc[-5]/10)    #Y1+Y2 
                else:
                    pc(5,(fc+(l-1)*2)).value=dc[-5]/10  #Y1
                    pc(6,(fc+(l-1)*2)).value=dc[-6]/10  #Y2
                
                if(dc[-4] == dc[-3]): #X1 = X2
                    merge(start_row = 7, start_column = fc+(l-1)*2, end_row = 8, end_column = fc+(l-1)*2)
                    cell = pc(5,(fc+(l-1)*2))
                    cell.value=2*(dc[-4]/10)    #X1+X2 
                else:
                    pc(7,(fc+(l-1)*2)).value=dc[-4]/10  #X1
                    pc(8,(fc+(l-1)*2)).value=dc[-3]/10  #X2
                
            else:            
                fc = 12     #fc: fator coluna = parametro inicial que indica a posição da coluna alvo

                if(dc[-5] == dc[-6]):   #Y1 == Y2
                    merge(start_row = 5, start_column = fc+(l-1)*2, end_row = 6, end_column = fc+(l-1)*2)
                    cell = pc(5,(fc+(l-1)*2))
                    cell.value=2*(dc[-5]/10)    #Y1+Y2 
                else:
                    pc(5,(fc+(l-1)*2)).value=dc[-5]/10  #Y1
                    pc(6,(fc+(l-1)*2)).value=dc[-6]/10  #Y2
                
                if(dc[-4] == dc[-3]):   #X1 = X2
                    merge(start_row = 7, start_column = fc+(l-1)*2, end_row = 8, end_column = fc+(l-1)*2)
                    cell = pc(5,(fc+(l-1)*2))
                    cell.value=2*(dc[-4]/10)    #X1+X2 
                else:
                    pc(7,(fc+(l-1)*2)).value=dc[-4]/10  #X1
                    pc(8,(fc+(l-1)*2)).value=dc[-3]/10  #X2                             
                
        elif((dc[6] and dc[7])==0 and (dc[8] or dc[9])!=0): #Y2 e  Y1 == 0
            Y2=dc[0]
            Y1=Y2
            if(l<3):
                fc = 9      #fc: fator coluna = parametro inicial que indica a posição da coluna alvo 
                merge(start_row = 5, start_column = fc+(l-1)*2, end_row = 6, end_column = fc+(l-1)*2)
                cell = pc(5,(fc+(l-1)*2))
                cell.value=Y2   #Y1 e Y2 = Y_merge 

                if(dc[-4] == dc[-3]): #X1 = X2
                    merge(start_row = 7, start_column = fc+(l-1)*2, end_row = 8, end_column = fc+(l-1)*2)
                    cell = pc(7,(fc+(l-1)*2))
                    cell.value=2*(dc[-4]/10)    #X1+X2                 
                else:
                    pc(7,(9+(l-1)*2)).value=dc[-4]/10   #X1
                    pc(8,(9+(l-1)*2)).value=dc[-3]/10   #X2
                
            elif(l<6):
                fc = 10     #fc: fator coluna = parametro inicial que indica a posição da coluna alvo
                merge(start_row = 5, start_column = fc+(l-1)*2, end_row = 6, end_column = fc+(l-1)*2)
                cell = pc(5,(fc+(l-1)*2))
                cell.value=Y2   #Y1 e Y2 = Y_merge 

                if(dc[-4] == dc[-3]): #X1 = X2
                    merge(start_row = 7, start_column = fc+(l-1)*2, end_row = 8, end_column = fc+(l-1)*2)
                    cell = pc(7,(fc+(l-1)*2))
                    cell.value=2*(dc[-4]/10)    #X1+X2                 
                else:
                    pc(7,(9+(l-1)*2)).value=dc[-4]/10   #X1
                    pc(8,(9+(l-1)*2)).value=dc[-3]/10   #X2

            else:
                fc = 12     #fc: fator coluna = parametro inicial que indica a posição da coluna alvo

                merge(start_row = 5, start_column = fc+(l-1)*2, end_row = 6, end_column = fc+(l-1)*2)
                cell = pc(5,(fc+(l-1)*2))
                cell.value=Y2   #Y1 e Y2 = Y_merge 
                
                if(dc[-4] == dc[-3]):   #X1 = X2
                    merge(start_row = 7, start_column = fc+(l-1)*2, end_row = 8, end_column = fc+(l-1)*2)
                    cell = pc(7,(fc+(l-1)*2))
                    cell.value=2*(dc[-4]/10)    #X1+X2                 
                else:
                    pc(7,(9+(l-1)*2)).value=dc[-4]/10   #X1
                    pc(8,(9+(l-1)*2)).value=dc[-3]/10   #X2
        
        elif((dc[8] and dc[9])==0 and (dc[6] or dc[7])!=0):     #X2 e X1 == 0
            X2=dc[1]
            X1=X2
            if(l<3):
                fc = 9      #fc: fator coluna = parametro inicial que indica a posição da coluna alvo
                merge(start_row = 7, start_column = fc+(l-1)*2, end_row = 8, end_column = fc+(l-1)*2)
                cell = pc(7,(fc+(l-1)*2))
                cell.value=X2       #X1 e X2 = X_merge 
                
                if(dc[-5] == dc[-6]):    #Y1 = Y2
                    merge(start_row = 5, start_column = fc+(l-1)*2, end_row = 6, end_column = fc+(l-1)*2)
                    cell = pc(5,(fc+(l-1)*2))
                    cell.value=2*(dc[-5]/10)    #Y1+Y2                 
                else:
                    pc(5,(fc+(l-1)*2)).value=dc[-5]/10  #Y1
                    pc(6,(fc+(l-1)*2)).value=dc[-6]/10  #Y2                               
                
            elif(l<6):
                fc = 10     #fc: fator coluna = parametro inicial que indica a posição da coluna alvo

                merge(start_row = 7, start_column = fc+(l-1)*2, end_row = 8, end_column = fc+(l-1)*2)
                cell = pc(7,(fc+(l-1)*2))
                cell.value=X2   #X1 e X2 = X_merge 
                
                if(dc[-5] == dc[-6]): #Y1 = Y2
                    merge(start_row = 5, start_column = fc+(l-1)*2, end_row = 6, end_column = fc+(l-1)*2)
                    cell = pc(5,(fc+(l-1)*2))
                    cell.value=2*(dc[-5]/10)    #Y1+Y2                 
                else:
                    pc(5,(fc+(l-1)*2)).value=dc[-5]/10  #Y1
                    pc(6,(fc+(l-1)*2)).value=dc[-6]/10  #Y2               
                
            else:
                fc = 12     #fc: fator coluna = parametro inicial que indica a posição da coluna alvo
                merge(start_row = 7, start_column = fc+(l-1)*2, end_row = 8, end_column = fc+(l-1)*2)
                cell = pc(7,(fc+(l-1)*2))
                cell.value=X2   #X1 e X2 = X_merge 
                
                if(dc[-5] == dc[-6]):   #Y1 = Y2
                    merge(start_row = 5, start_column = fc+(l-1)*2, end_row = 6, end_column = fc+(l-1)*2)
                    cell = pc(5,(fc+(l-1)*2))
                    cell.value=2*(dc[-5]/10)    #Y1+Y2                 
                else:
                    pc(5,(fc+(l-1)*2)).value=dc[-5]/10  #Y1
                    pc(6,(fc+(l-1)*2)).value=dc[-6]/10  #Y2    
    
        elif(dc[6]==0 and dc[7]==0 and dc[8]==0 and dc[9]==0):
            if(l<3):
                fc = 8  #fc: fator coluna = parametro inicial que indica a posição da coluna alvo
                pc(4,(fc+(l-1)*2)).value="{} x {}".format(dc[0],dc[1])  #Y x X
                
            elif(l<6):
                fc = 9  #fc: fator coluna = parametro inicial que indica a posição da coluna alvo
                pc(4,(fc+(l-1)*2)).value="{} x {}".format(dc[0],dc[1])  #Y e X                
                
            else:
                fc = 11     #fc: fator coluna = parametro inicial que indica a posição da coluna alvo
                pc(4,(fc+(l-1)*2)).value="{} x {}".format(dc[0],dc[1])  #Y e X
    

def setup3(arq1, dose_boost, primeiro_campo, total_campos, planilha):
    pc = planilha.cell
    merge = planilha.merge_cells
    pc(1, 4, value=dose(arq1))
    pc(1, 7, value=aplicacoes(arq1))
    pc(15,17).value = dose_boost    #Add dose incial do Boost Concomitante
    
    for l in range(primeiro_campo, primeiro_campo+total_campos):
        dados_campo = conf_campos(arq1,l)
        dc = dados_campo
        
        if(l<3):
            fc = 8  #fc: fator coluna = parametro inicial que indica a posição da coluna alvo
            pc(3,(fc+(l-1)*2)).value=dc[-2]     #Add Distancia fonte pele
            pc(9,(fc+(l-1)*2)).value=dc[2]      #Add Angulo do Gantry
            pc(10,(fc+(l-1)*2)).value=dc[3]     #Add Angulo do Colimador
            pc(13,(fc+(l-1)*2)).value=dc[-1]    #Add Unidade Monitora

            if(dc[4]!="NONE"):
                pc(11,(fc+(l-1)*2)).value="{} {}".format(dc[4], dc[5])  #Add Tipo e Posição do Filtro Fisico
            else:
                pc(11,(fc+(l-1)*2),value="")    #Add Tipo e Posição do Filtro Fisico
            
        elif(l<5):
            fc = 9  #fc: fator coluna = parametro inicial que indica a posição da coluna alvo
            pc(3,(fc+(l-1)*2)).value=dc[-2]     #Add Distancia fonte pele
            pc(9,(fc+(l-1)*2)).value=dc[2]      #Add Angulo do Gantry
            pc(10,(fc+(l-1)*2)).value=dc[3]     #Add Angulo do Colimador
            pc(13,(fc+(l-1)*2)).value=dc[-1]    #Add Unidade Monitora

            if(dc[4]!="NONE"):
                pc(11,(fc+(l-1)*2)).value="{} {}".format(dc[4], dc[5])  #Add Tipo e Posição do Filtro Fisico
            else:
                pc(11,(fc+(l-1)*2),value="")    #Add Tipo e Posição do Filtro Fisico
        
        else:
            fc = 11     #fc: fator coluna = parametro inicial que indica a posição da coluna alvo
            pc(3,(fc+(l-1)*2)).value=dc[-2]     #Add Distancia fonte pele
            pc(9,(fc+(l-1)*2)).value=dc[2]      #Add Angulo do Gantry
            pc(10,(fc+(l-1)*2)).value=dc[3]     #Add Angulo do Colimador
            pc(13,(fc+(l-1)*2)).value=dc[-1]    #Add Unidade Monitora

            if(dc[4]!="NONE"):
                pc(11,(fc+(l-1)*2)).value="{} {}".format(dc[4], dc[5])  #Add Tipo e Posição do Filtro Fisico
            else:
                pc(11,(fc+(l-1)*2),value="")    #Add Tipo e Posição do Filtro Fisico                
        
        #Condiçoes e Coordenadas para Y, X, Y2, Y1, X1, X2
                        
        if((dc[6] or dc[7])!=0 and (dc[8] or dc[9])!=0):
            if(l<3):
                fc = 9  #fc: fator coluna = parametro inicial que indica a posição da coluna alvo

                if(dc[-5] == dc[-6]): #Y1 == Y2
                    merge(start_row = 5, start_column = fc+(l-1)*2, end_row = 6, end_column = fc+(l-1)*2)
                    cell = pc(5,(fc+(l-1)*2))
                    cell.value = 2*(dc[-5]/10)  #Y1+Y2 
                else:
                    pc(5,(fc+(l-1)*2)).value=dc[-5]/10  #Y1
                    pc(6,(fc+(l-1)*2)).value=dc[-6]/10  #Y2
                
                if(dc[-4] == dc[-3]): #X1 = X2
                    merge(start_row = 7, start_column = fc+(l-1)*2, end_row = 8, end_column = fc+(l-1)*2)
                    cell = pc(5,(fc+(l-1)*2))
                    cell.value = 2*(dc[-4]/10)  #X1+X2 
                else:
                    pc(7,(fc+(l-1)*2)).value=dc[-4]/10  #X1
                    pc(8,(fc+(l-1)*2)).value=dc[-3]/10  #X2

            elif(l<5):
                fc = 10     #fc: fator coluna = parametro inicial que indica a posição da coluna alvo

                if(dc[-5] == dc[-6]): #Y1 == Y2
                    merge(start_row = 5, start_column = fc+(l-1)*2, end_row = 6, end_column = fc+(l-1)*2)
                    cell = pc(5,(fc+(l-1)*2))
                    cell.value = 2*(dc[-5]/10)  #Y1+Y2 
                else:
                    pc(5,(fc+(l-1)*2)).value=dc[-5]/10  #Y1
                    pc(6,(fc+(l-1)*2)).value=dc[-6]/10  #Y2
                
                if(dc[-4] == dc[-3]):   #X1 = X2
                    merge(start_row = 7, start_column = fc+(l-1)*2, end_row = 8, end_column = fc+(l-1)*2)
                    cell = pc(5,(fc+(l-1)*2))
                    cell.value = 2*(dc[-4]/10)  #X1+X2 
                else:
                    pc(7,(fc+(l-1)*2)).value=dc[-4]/10  #X1
                    pc(8,(fc+(l-1)*2)).value=dc[-3]/10  #X2

            else:            
                fc = 12 #fc: fator coluna = parametro inicial que indica a posição da coluna alvo

                if(dc[-5] == dc[-6]):   #Y1 == Y2
                    merge(start_row = 5, start_column = fc+(l-1)*2, end_row = 6, end_column = fc+(l-1)*2)
                    cell = pc(5,(fc+(l-1)*2))
                    cell.value = 2*(dc[-5]/10)  #Y1+Y2 
                else:
                    pc(5,(fc+(l-1)*2)).value=dc[-5]/10  #Y1
                    pc(6,(fc+(l-1)*2)).value=dc[-6]/10  #Y2
                
                if(dc[-4] == dc[-3]): #X1 = X2
                    merge(start_row = 7, start_column = fc+(l-1)*2, end_row = 8, end_column = fc+(l-1)*2)
                    cell = pc(5,(fc+(l-1)*2))
                    cell.value=2*(dc[-4]/10)    #X1+X2 
                else:
                    pc(7,(fc+(l-1)*2)).value=dc[-4]/10  #X1
                    pc(8,(fc+(l-1)*2)).value=dc[-3]/10  #X2 

        elif((dc[6] and dc[7])==0 and (dc[8] or dc[9])!=0): #Y2 e  Y1 == 0
            Y2=dc[0]
            Y1=Y2
            if(l<3):
                fc = 9  #fc: fator coluna = parametro inicial que indica a posição da coluna alvo
                merge(start_row = 5, start_column = fc+(l-1)*2, end_row = 6, end_column = fc+(l-1)*2)
                cell = pc(5,(fc+(l-1)*2))
                cell.value=Y2   #Y1 e Y2 = Y_merge 

                if(dc[-4] == dc[-3]): #X1 = X2
                    merge(start_row = 7, start_column = fc+(l-1)*2, end_row = 8, end_column = fc+(l-1)*2)
                    cell = pc(7,(fc+(l-1)*2))
                    cell.value=2*(dc[-4]/10)    #X1+X2                 
                else:
                    pc(7,(9+(l-1)*2)).value=dc[-4]/10   #X1
                    pc(8,(9+(l-1)*2)).value=dc[-3]/10   #X2

            elif(l<5):
                fc = 10     #fc: fator coluna = parametro inicial que indica a posição da coluna alvo
                merge(start_row = 5, start_column = fc+(l-1)*2, end_row = 6, end_column = fc+(l-1)*2)
                cell = pc(5,(fc+(l-1)*2))
                cell.value=Y2   #Y1 e Y2 = Y_merge        

                if(dc[-4] == dc[-3]): #X1 = X2
                    merge(start_row = 7, start_column = fc+(l-1)*2, end_row = 8, end_column = fc+(l-1)*2)
                    cell = pc(7,(fc+(l-1)*2))
                    cell.value=2*(dc[-4]/10)    #X1+X2                 
                else:
                    pc(7,(9+(l-1)*2)).value=dc[-4]/10   #X1
                    pc(8,(9+(l-1)*2)).value=dc[-3]/10   #X2

            else:
                fc = 12     #fc: fator coluna = parametro inicial que indica a posição da coluna alvo
                merge(start_row = 5, start_column = fc+(l-1)*2, end_row = 6, end_column = fc+(l-1)*2)
                cell = pc(5,(fc+(l-1)*2))
                cell.value=Y2   #Y1 e Y2 = Y_merge 
                
                if(dc[-4] == dc[-3]): #X1 = X2
                    merge(start_row = 7, start_column = fc+(l-1)*2, end_row = 8, end_column = fc+(l-1)*2)
                    cell = pc(7,(fc+(l-1)*2))
                    cell.value=2*(dc[-4]/10)    #X1+X2                 
                else:
                    pc(7,(9+(l-1)*2)).value=dc[-4]/10   #X1
                    pc(8,(9+(l-1)*2)).value=dc[-3]/10   #X2

        elif((dc[8] and dc[9])==0 and (dc[6] or dc[7])!=0): #X2 e X1 == 0
            X2=dc[1]
            X1=X2
            if(l<3):
                fc = 9 #fc: fator coluna = parametro inicial que indica a posição da coluna alvo
                merge(start_row = 7, start_column = fc+(l-1)*2, end_row = 8, end_column = fc+(l-1)*2)
                cell = pc(7,(fc+(l-1)*2))
                cell.value=X2   #X1 e X2 = X_merge 
                
                if(dc[-5] == dc[-6]):   #Y1 = Y2
                    merge(start_row = 5, start_column = fc+(l-1)*2, end_row = 6, end_column = fc+(l-1)*2)
                    cell = pc(5,(fc+(l-1)*2))
                    cell.value=2*(dc[-5]/10)    #Y1+Y2                 
                else:
                    pc(5,(fc+(l-1)*2)).value=dc[-5]/10  #Y1
                    pc(6,(fc+(l-1)*2)).value=dc[-6]/10  #Y2 

            elif(l<5):
                fc = 10     #fc: fator coluna = parametro inicial que indica a posição da coluna alvo
                merge(start_row = 7, start_column = fc+(l-1)*2, end_row = 8, end_column = fc+(l-1)*2)
                cell = pc(7,(fc+(l-1)*2))
                cell.value=X2   #X1 e X2 = X_merge 
                
                if(dc[-5] == dc[-6]): #Y1 = Y2
                    merge(start_row = 5, start_column = fc+(l-1)*2, end_row = 6, end_column = fc+(l-1)*2)
                    cell = pc(5,(fc+(l-1)*2))
                    cell.value=2*(dc[-5]/10)    #Y1+Y2                 
                else:
                    pc(5,(fc+(l-1)*2)).value=dc[-5]/10  #Y1
                    pc(6,(fc+(l-1)*2)).value=dc[-6]/10  #Y2      
                
            else:
                fc = 12 #fc: fator coluna = parametro inicial que indica a posição da coluna alvo
                merge(start_row = 7, start_column = fc+(l-1)*2, end_row = 8, end_column = fc+(l-1)*2)
                cell = pc(7,(fc+(l-1)*2))
                cell.value=X2   #X1 e X2 = X_merge 
                
                if(dc[-5] == dc[-6]): #Y1 = Y2
                    merge(start_row = 5, start_column = fc+(l-1)*2, end_row = 6, end_column = fc+(l-1)*2)
                    cell = pc(5,(fc+(l-1)*2))
                    cell.value=2*(dc[-5]/10)    #Y1+Y2                 
                else:
                    pc(5,(fc+(l-1)*2)).value=dc[-5]/10  #Y1
                    pc(6,(fc+(l-1)*2)).value=dc[-6]/10  #Y2    

        elif(dc[6]==0 and dc[7]==0 and dc[8]==0 and dc[9]==0):
            if(l<3):
                fc = 8 #fc: fator coluna = parametro inicial que indica a posição da coluna alvo
                pc(4,(fc+(l-1)*2)).value="{} x {}".format(dc[0],dc[1]) #Y x X
                
            elif(l<5):
                fc = 9 #fc: fator coluna = parametro inicial que indica a posição da coluna alvo
                pc(4,(fc+(l-1)*2)).value="{} x {}".format(dc[0],dc[1])  #Y e X                
                
            else:
                fc = 11 #fc: fator coluna = parametro inicial que indica a posição da coluna alvo
                pc(4,(fc+(l-1)*2)).value="{} x {}".format(dc[0],dc[1])  #Y e X

def setup4(arq1, dose_boost, primeiro_campo, total_campos, planilha):
    pc = planilha.cell
    merge = planilha.merge_cells
    pc(1, 4, value=dose(arq1))
    pc(1, 7, value=aplicacoes(arq1))
    pc(15,17).value = dose_boost    #Add dose incial do Boost Concomitante
    
    for l in range(primeiro_campo, primeiro_campo+total_campos):
        dados_campo = conf_campos(arq1,l)
        dc = dados_campo
        
        if(l<3):
            fc = 8  #fc: fator coluna = parametro inicial que indica a posição da coluna alvo
        else:
            fc = 15

        pc(3,(fc+(l-1)*2)).value=dc[-2]     #Add Distancia fonte pele
        pc(9,(fc+(l-1)*2)).value=dc[2]      #Add Angulo do Gantry
        pc(10,(fc+(l-1)*2)).value=dc[3]     #Add Angulo do Colimador
        pc(13,(fc+(l-1)*2)).value=dc[-1]    #Add Unidade Monitora
        
        if(dc[4]!="NONE"):
            pc(11,(fc+(l-1)*2)).value="{} {}".format(dc[4], dc[5])  #Add Tipo e Posição do Filtro Fisico
        else:
            pc(11,(fc+(l-1)*2),value="")    #Add Tipo e Posição do Filtro Fisico            
        
        #Condiçoes e Coordenadas para Y, X, Y2, Y1, X1, X2
                        
        if((dc[6] or dc[7])!=0 and (dc[8] or dc[9])!=0):

            if(l<3):
                fc = 9 #fc: fator coluna = parametro inicial que indica a posição da coluna alvo
            else:
                fc = 16
            
            if(dc[-5] == dc[-6]): #Y1 == Y2
                merge(start_row = 5, start_column = fc+(l-1)*2, end_row = 6, end_column = fc+(l-1)*2)
                cell = pc(5,(fc+(l-1)*2))
                cell.value = 2*(dc[-5]/10)  #Y1+Y2 
            else:
                pc(5,(fc+(l-1)*2)).value=dc[-5]/10  #Y1
                pc(6,(fc+(l-1)*2)).value=dc[-6]/10  #Y2
            
            if(dc[-4] == dc[-3]): #X1 = X2
                merge(start_row = 7, start_column = fc+(l-1)*2, end_row = 8, end_column = fc+(l-1)*2)
                cell = pc(5,(fc+(l-1)*2))
                cell.value = 2*(dc[-4]/10)  #X1+X2 
            else:
                pc(7,(fc+(l-1)*2)).value=dc[-4]/10  #X1
                pc(8,(fc+(l-1)*2)).value=dc[-3]/10  #X2
                
        elif((dc[6] and dc[7])==0 and (dc[8] or dc[9])!=0): #Y2 e  Y1 == 0
            Y2=dc[0]
            Y1=Y2

            if(l<3):
                fc = 9 #fc: fator coluna = parametro inicial que indica a posição da coluna alvo
            else:
                fc = 16            
            merge(start_row = 5, start_column = fc+(l-1)*2, end_row = 6, end_column = fc+(l-1)*2)
            cell = pc(5,(fc+(l-1)*2))
            cell.value=Y2   #Y1 e Y2 = Y_merge 

            if(dc[-4] == dc[-3]): #X1 = X2
                merge(start_row = 7, start_column = fc+(l-1)*2, end_row = 8, end_column = fc+(l-1)*2)
                cell = pc(7,(fc+(l-1)*2))
                cell.value=2*(dc[-4]/10)    #X1+X2                 
            else:
                pc(7,(9+(l-1)*2)).value=dc[-4]/10   #X1
                pc(8,(9+(l-1)*2)).value=dc[-3]/10   #X2

        elif((dc[8] and dc[9])==0 and (dc[6] or dc[7])!=0): #X2 e X1 == 0
            X2=dc[1]
            X1=X2

            if(l<3):
                fc = 9 #fc: fator coluna = parametro inicial que indica a posição da coluna alvo
            else:
                fc = 16   
            merge(start_row = 7, start_column = fc+(l-1)*2, end_row = 8, end_column = fc+(l-1)*2)
            cell = pc(7,(fc+(l-1)*2))
            cell.value=X2   #X1 e X2 = X_merge 
            
            if(dc[-5] == dc[-6]): #Y1 = Y2
                merge(start_row = 5, start_column = fc+(l-1)*2, end_row = 6, end_column = fc+(l-1)*2)
                cell = pc(5,(fc+(l-1)*2))
                cell.value=2*(dc[-5]/10)    #Y1+Y2                 
            else:
                pc(5,(fc+(l-1)*2)).value=dc[-5]/10  #Y1
                pc(6,(fc+(l-1)*2)).value=dc[-6]/10  #Y2 

        elif(dc[6]==0 and dc[7]==0 and dc[8]==0 and dc[9]==0):
            if(l<3):
                fc = 8 #fc: fator coluna = parametro inicial que indica a posição da coluna alvo
            else:
                fc = 15    

            pc(4,(fc+(l-1)*2)).value="{} x {}".format(dc[0],dc[1])  #Y x X

#-----------------------------------/ Run Writer /------------------------------------------------------                
turno = 1

while True:
    make_dir_ficha()
    
    #-----------------/Escolher Modalidade da Planilha e Input arquivo PPF /---------------------        
    n, lista_arquivo = window_main(lista_arquivo)
    #---------------------------/Input Dose Boost/---------------------
    print(lista_arquivo)
    print(boost)
    if(n==3 or n == 4 or n == 5):
        dose_boost = window_boost()
    
    template = "TEMP{}.xlsx".format(n)

    excel = load_workbook(template)
    planilha = excel['MATRIZ']    
        
    #-------------------------//-------------------------
    only_file_name = []

    #--------------------/ Escolha Arquivo Principal para Dose /-----------------
    print(n)
    print(lista_arquivo)
    if(n == 3 or n == 4 or n == 5):
        lista_arquivo = sort_ppf(lista_arquivo)
        for arquivo in lista_arquivo:
            only_file_name.append(nome_arq(arquivo))

    print(lista_arquivo)
    

    #--------------------/ Preenchimeto Planilha_Template /----------------------
    print("\n\n------------------------")
    
    for arquivo_input in lista_arquivo:
        arq1 = arquivo_input#"./PPF/{}".format(name)
        
        if(arq1) == "Erro":
            messagebox.showinfo("Erro", message= "Erro ao adicionar os arquivos")
            break
        else:
            total_campos = quant_campos(arq1)
            primeiro_campo = campo_inicial(arq1)
            
            if(n == 1 or n == 2):
                setup1(arq1, primeiro_campo, total_campos, planilha)
            elif(n == 3):
                setup2(arq1, dose_boost, primeiro_campo, total_campos, planilha)
            elif(n == 4):
                setup3(arq1, dose_boost, primeiro_campo, total_campos, planilha)
            elif(n == 5):
                setup4(arq1, dose_boost, primeiro_campo, total_campos, planilha)
    
    name = only_file_name[0].split('.PPF')
    name = name[0]        
    excel.save('./Ficha_Tratamento/{}.xlsx'.format(name))

            #------------------------/ Continue Run /-------------------------------------
    Tk().withdraw()
    messagebox.showinfo("Concluído", message= "Ficha finalizada com sucesso!")
    break
